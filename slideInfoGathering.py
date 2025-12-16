import openpyxl as px

import constants as con
from constants import NUMBER_OF_DRIVERS, NUMBER_OF_PLAYERS, NUMBER_OF_TEAMS

class H2HReader():

    def __init__(self, sheet):

        self.sheet = sheet

    def gather_data(self):

        data_cols = 5
        col_offset = 2

        data_display = range(data_cols)
        player_preds = range(data_cols, data_cols + NUMBER_OF_PLAYERS)
        player_scores = range(data_cols + NUMBER_OF_PLAYERS, data_cols + 2 * NUMBER_OF_PLAYERS)

        data_rows = []
        player_pred_rows = []
        player_score_rows = []

        for i in range(col_offset, NUMBER_OF_TEAMS + col_offset):
            data_row = []
            player_pred_row = []
            player_score_row = []
            for ii in data_display:
                data_cell = self.sheet.cell(column=ii + 1, row=i + 1)
                data_row.append(data_cell.internal_value)
            for jj in player_preds:
                player_cell = self.sheet.cell(column=jj+1, row=i+1)
                player_pred_row.append(player_cell.internal_value)
            for kk in player_scores:
                player_cell = self.sheet.cell(column=kk+1, row=i+1)
                player_score_row.append(player_cell.internal_value)

            data_rows.append(data_row)
            player_pred_rows.append(player_pred_row)
            player_score_rows.append(player_score_row)

        self.data_rows = data_rows
        self.player_pred_rows = player_pred_rows
        self.player_score_rows = player_score_rows

    def format_to_slide(self):
        #Slides take:
        # drivers: tuple[str, str],
        # score: tuple[int, int],
        # player_preds: dict[str, str]
        # partial_scores = dict[str, int]

        players = sorted(con.PLAYERS_TO_COLOURS.keys())
        slides_data = []

        for i in range(len(self.data_rows)):
            data_row = self.data_rows[i]

            pairing = (data_row[0], data_row[1])
            score = (data_row[2], data_row[3])

            player_pred_row = self.player_pred_rows[i]
            player_preds = {}

            for j in range(len(players)):
                player = players[j]
                player_preds[player] = player_pred_row[j]

            slides_data.append((pairing, score, player_preds))

        return slides_data

    def get_scores(self):

        players = sorted(con.PLAYERS_TO_COLOURS.keys())
        player_scores = {}
        for i in range(len(con.PLAYERS_TO_COLOURS)):
            player = players[i]
            player_scores[player] = [sum([self.player_score_rows[k][i] for k in range(0, j+1)]) for j in range(len(self.player_score_rows))]
        self.player_scores = player_scores

        return player_scores

    def get_final_scores(self):
        final_scores = {player: scores[-1] for player, scores in self.player_scores.items()}
        return final_scores

class CCReader():
    def __init__(self, sheet):
        self.sheet = sheet

    def gather_data(self):
        teams = [self.sheet.cell(i + 2, column = 1).internal_value for i in range(1, NUMBER_OF_TEAMS + 1)]
        actual_order = [self.sheet.cell(i + 2, column = 2).internal_value for i in range(1, NUMBER_OF_TEAMS + 1)]
        player_orders = {self.sheet.cell(2,column =j+2).internal_value:
            [self.sheet.cell(i+2,column = j+2).internal_value for i in range(1, NUMBER_OF_TEAMS + 1)]
            for j in range(1, NUMBER_OF_PLAYERS+1)}
        player_scores = {self.sheet.cell(2, column = j).internal_value:
            self.sheet.cell(13, column = j).internal_value for j in range(10, 10+NUMBER_OF_PLAYERS)}

        self.teams_real_order = {teams[i]: actual_order[i] for i in range(NUMBER_OF_TEAMS)}
        self.player_orders = {player: {teams[j]: player_orders[player][j] for j in range(NUMBER_OF_TEAMS)} for player in player_orders}
        self.player_scores = player_scores

    def format_to_slide(self):
        return (self.teams_real_order, self.player_orders, self.player_scores)

    def get_scores(self):
        return self.player_scores

class DCReader():
    def __init__(self, sheet):
        self.sheet = sheet

    def gather_data(self):
        drivers = [self.sheet.cell(i + 2, column = 1).internal_value for i in range(1, NUMBER_OF_DRIVERS + 1)]
        actual_order = [self.sheet.cell(i + 2, column = 2).internal_value for i in range(1, NUMBER_OF_DRIVERS + 1)]
        player_orders = {self.sheet.cell(2, column = j+2).internal_value:
            [self.sheet.cell(i + 2, column = j+2).internal_value for i in range(1, NUMBER_OF_DRIVERS + 1)]
            for j in range(1, NUMBER_OF_PLAYERS+1)}
        player_scores = {self.sheet.cell(2, column = j).internal_value:
            self.sheet.cell(23, column = j).internal_value for j in range(10, 10+NUMBER_OF_PLAYERS)}

        self.driver_real_order = {drivers[i]: actual_order[i] for i in range(NUMBER_OF_DRIVERS)}
        self.player_orders = {player: {drivers[j]: player_orders[player][j] for j in range(NUMBER_OF_DRIVERS)} for player in player_orders}
        self.player_scores = player_scores

    def format_to_slide(self):
        return (self.driver_real_order, self.player_orders, self.player_scores)

    def get_scores(self):
        return self.player_scores

class TrueFalseReader():
    def __init__(self, sheet, shift = 7):
        self.sheet = sheet
        self.shift = shift

    def gather_data(self):
        self.driver_to_cond = {self.sheet.cell(i + 2, column = 1).internal_value: (self.sheet.cell(i + 2, column = 2).internal_value, self.sheet.cell(i + 2, column = 3).internal_value) for i in range(1, NUMBER_OF_DRIVERS + 1)}
        self.player_to_truths = {
            self.sheet.cell(2, column = 4 + j).internal_value:
            [self.sheet.cell(i+2, column = 1).internal_value
             for i in range(1, NUMBER_OF_DRIVERS+1)
             if self.sheet.cell(i+2, column = 4 + j).internal_value]
            for j in range(NUMBER_OF_PLAYERS)}
        self.player_scores = {self.sheet.cell(31 + i, 1).internal_value: self.sheet.cell(31 + i, 2).internal_value for i in range(NUMBER_OF_PLAYERS)}

    def format_to_slide(self):
        return (self.driver_to_cond, self.player_to_truths, self.player_scores)

    def get_scores(self):
        return self.player_scores

class SuperlativeDriverReader():
    def __init__(
            self,
            sheet,
            driver_col = 1,
            metric_col = 2,
            rank_col = 3,
            player_col = 30,
            player_row = 4
        ):
        self.sheet = sheet
        self.driver_col = driver_col
        self.metric_col = metric_col
        self.rank_col = rank_col
        self.player_col = player_col
        self.player_row = player_row

    def gather_data(self):
        self.driver_ranking = [
            (self.sheet.cell(i, self.rank_col).internal_value,
            self.sheet.cell(i, self.driver_col).internal_value,
            self.sheet.cell(i, self.metric_col).internal_value)
            for i in range(3, 3 + NUMBER_OF_DRIVERS)
            if self.sheet.cell(i, self.rank_col).internal_value < 10.5
        ]
        self.driver_ranking.sort(key = lambda x: float(x[0]))

        self.player_preds = {
            self.sheet.cell(self.player_row + i, self.player_col).internal_value:
            (self.sheet.cell(self.player_row + i, self.player_col + 1).internal_value,
            self.sheet.cell(self.player_row + i, self.player_col + 3).internal_value)
            for i in range(NUMBER_OF_PLAYERS)
        }

        self.player_score = {
            self.sheet.cell(self.player_row + i, self.player_col).internal_value:
            self.sheet.cell(self.player_row + i, self.player_col + 3).internal_value
            for i in range(NUMBER_OF_PLAYERS)
        }

    def format_to_slide(self):
        return self.driver_ranking, self.player_preds

    def get_scores(self):
        return self.player_score

class SuperlativeTeamReader():
    def __init__(
            self,
            sheet,
            team_col = 10,
            metric_col = 11,
            rank_col = 12,
            player_col = 13,
            prediction_col = 14,
            score_col = 16,
            team_row_delta = 2,
            team_row = 2,
            player_row = 2
        ):
        self.sheet = sheet
        self.team_col = team_col
        self.metric_col = metric_col
        self.rank_col = rank_col
        self.player_col = player_col
        self.prediction_col = prediction_col
        self.score_col = score_col
        self.team_row_delta = team_row_delta
        self.team_row = team_row
        self.player_row = player_row

    def gather_data(self):
        self.team_ranking = sorted([
            (self.sheet.cell(i, self.rank_col).internal_value,
            self.sheet.cell(i, self.team_col).internal_value,
            self.sheet.cell(i, self.metric_col).internal_value)
            for i in range(self.team_row, self.team_row + self.team_row_delta * NUMBER_OF_TEAMS, self.team_row_delta)
        ],
        key = lambda x: float(x[0]))

        self.player_preds = {
            self.sheet.cell(self.player_row + i, self.player_col).internal_value:
            (self.sheet.cell(self.player_row + i, self.prediction_col).internal_value,
            self.sheet.cell(self.player_row + i, self.score_col).internal_value)
            for i in range(NUMBER_OF_PLAYERS)
        }

        self.player_score = {
            self.sheet.cell(self.player_row + i, self.player_col).internal_value:
            self.sheet.cell(self.player_row + i, self.score_col).internal_value
            for i in range(NUMBER_OF_PLAYERS)
        }

    def format_to_slide(self):
        return self.team_ranking, self.player_preds

    def get_scores(self):
        return self.player_score

class First6RacesReader():
    def __init__(self, wb):
        self.n_after_n = wb

    def gather_data(self):
        n_after_n = self.n_after_n
        self.orders = {n_after_n.cell(2 + i, 1).internal_value: [n_after_n.cell(2 + i, j+2).internal_value for j in range(6)] for i in range(NUMBER_OF_DRIVERS)}
        self.player_preds = {n_after_n.cell(2 + i * 3, 9).internal_value:
                [(n_after_n.cell(2 + i * 3, 10 + j).internal_value,
                n_after_n.cell(4 + i * 3, 10 + j).internal_value)
                for j in range(6)]
                for i in range(NUMBER_OF_PLAYERS)}
        self.player_scores = {n_after_n.cell(31 + i, 1).internal_value:
                              n_after_n.cell(31 + i, 3).internal_value
                              for i in range(NUMBER_OF_PLAYERS)}

    def format_to_slide(self):
        return (self.orders, self.player_preds)

    def get_scores(self):
        return self.player_scores

class Pick5RacesReader():
    def __init__(self, wb):
        self.wb = wb

    def gather_data(self):
        pick5 = self.wb
        self.player_preds = {
            pick5.cell(2 + 2 * i, 1).internal_value:
            [(pick5.cell(2 + 2 * i, 2 + j).internal_value,
            pick5.cell(3 + 2 * i, 2 + j).internal_value if pick5.cell(3 + 2 * i, 2 + j).internal_value else 0)
            for j in range(5)]
            for i in range(NUMBER_OF_PLAYERS)
        }
        self.all_races = set()
        for pred_list in self.player_preds.values():
            for race, score in pred_list:
                self.all_races.add(race)
        self.player_scores = {pick5.cell(2 + 2 * i, 1).internal_value: pick5.cell(3 + 2 * i, 1).internal_value for i in range(NUMBER_OF_PLAYERS)}

    def format_to_slide(self):
        return (self.player_preds, self.all_races)

    def get_scores(self):
        return self.player_scores